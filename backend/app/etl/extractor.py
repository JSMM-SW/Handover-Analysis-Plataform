"""Extracción de información básica de un archivo Excel.

Alcance de esta iteración: identificar hojas y su forma (filas, columnas,
encabezados), sin interpretar el significado de las columnas. Las reglas de
qué hoja/columnas son relevantes para el dominio de handover se definirán
tras analizar un archivo real (Objetivo 2 del plan de tesis).
"""

import logging
from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from app.core.exceptions import ExtractionError, SchemaValidationError
from app.etl.constants import REQUIRED_COLUMNS
from app.schemas.ingestion import SheetInfo

logger = logging.getLogger(__name__)

_INVALID_EXCEL_MESSAGE = (
    "El archivo no pudo ser leído como un Excel válido (formato incompatible o corrupto)."
)


def _open_workbook(path: Path, execution_id: str):
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except (InvalidFileException, OSError, KeyError) as exc:
        logger.warning("[%s] No se pudo abrir el archivo como Excel: %s", execution_id, exc)
        raise ExtractionError(_INVALID_EXCEL_MESSAGE) from exc
    except Exception as exc:  # zipfile.BadZipFile y otros errores internos de openpyxl
        logger.warning("[%s] Error inesperado al abrir el archivo: %s", execution_id, exc)
        raise ExtractionError(_INVALID_EXCEL_MESSAGE) from exc


def extract_basic_info(path: Path, execution_id: str) -> list[SheetInfo]:
    workbook = _open_workbook(path, execution_id)

    try:
        if not workbook.sheetnames:
            raise ExtractionError("El archivo no contiene hojas.")

        sheets: list[SheetInfo] = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            header_row = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ()
            )
            headers = ["" if value is None else str(value) for value in header_row]

            sheets.append(
                SheetInfo(
                    name=sheet_name,
                    num_rows=worksheet.max_row or 0,
                    num_cols=worksheet.max_column or 0,
                    headers=headers,
                )
            )
        return sheets
    finally:
        workbook.close()


def validate_required_columns(sheets: list[SheetInfo], execution_id: str) -> None:
    """Verifica que cada hoja tenga las columnas de negocio requeridas.

    Es un chequeo estructural (bloquea todo el archivo si falla), distinto
    de las validaciones por registro que hace el Validator sobre cada fila.
    """
    missing_by_sheet: dict[str, list[str]] = {}
    for sheet in sheets:
        missing = [col for col in REQUIRED_COLUMNS if col not in sheet.headers]
        if missing:
            missing_by_sheet[sheet.name] = missing

    if missing_by_sheet:
        details = "; ".join(
            f"'{sheet}' no tiene: {', '.join(cols)}"
            for sheet, cols in missing_by_sheet.items()
        )
        logger.warning("[%s] Columnas requeridas ausentes: %s", execution_id, details)
        raise SchemaValidationError(
            f"El archivo no tiene la estructura esperada. {details}."
        )


def extract_records(path: Path, execution_id: str) -> list[dict]:
    """Lee todas las hojas del Excel y devuelve las filas de negocio crudas.

    Cada elemento es {"hoja_origen": str, "fila_excel": int, "data": {...}},
    donde "data" solo contiene REQUIRED_COLUMNS con sus valores tal como
    vienen del Excel (sin limpiar ni normalizar todavía).
    """
    workbook = _open_workbook(path, execution_id)

    try:
        records: list[dict] = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            header_row = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ()
            )
            headers = ["" if value is None else str(value) for value in header_row]
            column_index = {col: headers.index(col) for col in REQUIRED_COLUMNS if col in headers}

            for excel_row_number, row in enumerate(
                worksheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                if row is None or all(value is None for value in row):
                    continue  # fila completamente vacía, no es un registro
                data = {col: row[idx] for col, idx in column_index.items()}
                records.append(
                    {
                        "hoja_origen": sheet_name,
                        "fila_excel": excel_row_number,
                        "data": data,
                    }
                )

        logger.info("[%s] Extracción completa: %d registros crudos", execution_id, len(records))
        return records
    finally:
        workbook.close()
