"""Extracción de información básica de un archivo Excel.

Alcance de esta iteración: identificar hojas y su forma (filas, columnas,
encabezados), sin interpretar el significado de las columnas. Las reglas de
qué hoja/columnas son relevantes para el dominio de handover se definirán
tras analizar un archivo real (Objetivo 2 del plan de tesis).
"""

import logging
from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from app.shared.exceptions import ExtractionError
from app.modules.ingesta.schemas import SheetInfo

logger = logging.getLogger(__name__)


def extract_basic_info(path: Path, execution_id: str) -> list[SheetInfo]:
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except (InvalidFileException, OSError, KeyError) as exc:
        logger.warning("[%s] No se pudo abrir el archivo como Excel: %s", execution_id, exc)
        raise ExtractionError(
            "El archivo no pudo ser leído como un Excel válido (formato incompatible o corrupto)."
        ) from exc
    except Exception as exc:  # zipfile.BadZipFile y otros errores internos de openpyxl
        logger.warning("[%s] Error inesperado al abrir el archivo: %s", execution_id, exc)
        raise ExtractionError(
            "El archivo no pudo ser leído como un Excel válido (formato incompatible o corrupto)."
        ) from exc

    try:
        if not workbook.sheetnames:
            raise ExtractionError("El archivo no contiene hojas.")

        sheets: list[SheetInfo] = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            header_row = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ()
            )
            headers = ["" if value is None else str(value) for value in header_row]

            sheets.append(
                SheetInfo(
                    name=sheet_name,
                    num_rows=worksheet.max_row or 0,
                    num_cols=worksheet.max_column or 0,
                    headers=headers,
                )
            )
        return sheets
    finally:
        workbook.close()
